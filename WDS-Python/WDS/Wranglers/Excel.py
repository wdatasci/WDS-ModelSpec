'''Extract Excel worksheets to CSV files..

Copyright 2019, Wypasek Data Science, Inc., (WDataSci/WDS)
'''

import os,sys
import pudb
import argparse
import traceback

import collections

import time
import os.path as osp
import copy
import glob
import fnmatch
import re

import math
from datetime import date

#old, for some reason xlrd in python3.8 will not read xlsm, duh
##using xlrd for reading, but will use xlsxwriter instead of xlwt
import xlrd
#new-er openpyxl might have a problem with loading too much into memory
from openpyxl import load_workbook as xl
import csv


import WDS.Wranglers.dir_walk
import WDS.Util.MonthID as mMonthID


def ExcelLayout(fn):
        
        w=xlrd.open_workbook(fn)

        print('Excel Layout of ', fn)
        print()
        print('  Sheet Names')
        for nm in w.sheet_names():
            print('    ',nm)

        print()
        print('  Named ranges')
        for k in w.name_map:
            r=w.name_map[k]
            if type(r) is list:
                r=r[0]
            print('    name=%s, scope=%d, formula_text=%s' % (r.name,r.scope,r.formula_text))

        print()
        print('datemode of workbook:',w.datemode)
        print()
        print('dir of python workbook object')
        print(dir(w))

def Excel2CSV_NameTransform_Example(lName,lFullPath):
    '''Excel2CSV, if given a NameTransform named argument, 
        expects a function with two arguments.  It is passed
        the sheet specific WBName+"."+SheetPrefix+"."+WSName
        and the original name of the file (which can include
        the path).  The NameTransform argument can then
        employ logic around both.'''
    rv=lName.strip()
    return rv


def Excel2CSV(fn
            , targetdir="."
            , targetbasename=None
            , SheetPrefix="Sheet"
            , NameTransform=None
            , nrows=None
            , ncols=None
            , isPreviewNameOnly=False
            ):
    if not osp.isdir(targetdir):
        os.makedirs(targetdir,mode=777,exist_ok=True)
        #osp.makedirs(targetdir,mode=777,exist_ok=True)
    fnh,fnt = osp.split(fn)
    fntr,fnte = osp.splitext(fnt)
    #old
    #w=xlrd.open_workbook(fn)
    w=xl(filename=fn,data_only=True) #read_only=True,data_only=True)
    w_base_date=w.epoch
    for sname in w.sheetnames:
        newfn=copy.copy(sname)
        s=w[sname]
        if targetbasename:
            newfn=targetbasename+"."+SheetPrefix+"."+newfn+".csv"
        else:
            newfn=fntr+"."+SheetPrefix+"."+newfn+".csv"
        if NameTransform:
            newfn=NameTransform(newfn,fn)
        #old
        #l_nrows=nrows if nrows is not None else s.nrows
        #l_ncols=ncols if ncols is not None else s.ncols
        l_nrows=nrows if nrows is not None else s.max_row
        l_ncols=ncols if ncols is not None else s.max_column
        if isPreviewNameOnly:
            print("Original:",fn," Sheet:",sname)
            print("  Extract To:",newfn)
        else:
            fid=open(osp.join(targetdir,newfn),'w')
            dw=csv.DictWriter(fid, dialect=csv.excel, fieldnames=list(range(l_ncols)))
            for i in range(1,l_nrows+1):
                lrow=collections.OrderedDict()
                for j in range(1,l_ncols+1):
                    jM1=j-1
                    #if (s.cell(i,j).ctype==1): 
                    if (s.cell(i,j).data_type=='s'): 
                        lrow[jM1]=s.cell(i,j).value.strip()
                    elif (s.cell(i,j).data_type=='d' or s.cell(i,j).is_date): 
                        dte=mMonthID.CleanDate(s.cell(i,j).value)
                        lrow[jM1]=mMonthID.Date2isoformat(dte)
                        if 0:
                            if (math.fabs(math.trunc(s.cell(i,j).value)-s.cell(i,j).value)<1e-8):
                                y,m,d,hh,mm,ss=xlrd.xldate.xldate_as_tuple(s.cell(i,j).value,w.datemode)
                                lrow[jM1]=str(date(y,m,d)).strip()
                            else:
                                lrow[jM1]=str(xlrd.xldate.xldate_as_datetime(s.cell(i,j).value,w.datemode)).strip()
                    elif (s.cell(i,j).data_type=='b'): 
                        lrow[jM1]=s.cell(i,j).value
                    elif (s.cell(i,j).data_type=='n'): 
                        if s.cell(i,j).value is None:
                            lrow[jM1]=None
                        elif (math.fabs(math.trunc(s.cell(i,j).value)-s.cell(i,j).value)<1e-8):
                            lrow[jM1]=int(s.cell(i,j).value)
                        else:
                            lrow[jM1]=s.cell(i,j).value
                dw.writerow(lrow)
            fid.close()


if __name__=="__main__":
    def main_argparser():
        _parser=argparse.ArgumentParser()
        #positional
        #default positional
        fn='./zzzExamples/data/Book1.xlsm'
        _parser.add_argument("Excel_to_load"
                            , help="first argument, xml to load"
                            , nargs='?'
                            , default=fn
                            )
        #optional switches
        _parser.add_argument("--pudb"
                            , help="turn on the pudb debugger before main"
                            , action="store_true"
                            )
        return _parser


    def main(args=None):

        ExcelLayout(args.Excel_to_load)
        Excel2CSV(args.Excel_to_load
                    , targetdir="./gtmp"
                    )
        
    #end def main
        

    l_argparser = main_argparser()
    try:
        args=l_argparser.parse_args()
        if args.pudb:
            pudb.set_trace()
        main(args=args)
    except Exception as e:
        print("Hey")
        print(e)
        print(traceback.format_tb(e.__traceback__))






